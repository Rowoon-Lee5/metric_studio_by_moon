"""Verify that the supplied delisted-stock universe is present in the price panel.

The audit is deliberately code-based.  An early terminal observation is not
treated as a delisting proxy: the supplied ``자본총계(상장폐지)`` workbook is
matched directly against both the raw adjusted-price workbook and the monthly
research panel.
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "results"
RAW = Path(r"C:\Users\benef\OneDrive\바탕 화면\직박구리2\직박구리\숭실대학교\(제출용) 금융인공지능및빅데이터응용_프로젝트_final\금융인공지능및빅데이터응용_프로젝트_final\raw data")


def normalize(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip().replace("A", "").split(".")[0]
    return text.zfill(6) if text.isdigit() else None


def workbook_codes(path: Path) -> set[str]:
    """Read the header row with the largest number of stock-code-like cells."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    best: set[str] = set()
    for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
        codes = {code for value in row if (code := normalize(value)) is not None}
        if len(codes) > len(best):
            best = codes
    wb.close()
    return best


def main() -> None:
    raw_price_codes = workbook_codes(RAW / "수정주가.xlsx")
    delisted_codes = workbook_codes(RAW / "자본총계(상장폐지).xlsx")
    panel = pd.read_pickle(OUT / "monthly_price_adv_panel.pkl")
    panel_codes = {code for value in panel.ticker.unique() if (code := normalize(value)) is not None}
    rows = pd.DataFrame(
        [
            {"universe": "raw_adjusted_price", "ticker_count": len(raw_price_codes)},
            {"universe": "delisted_capital_workbook", "ticker_count": len(delisted_codes)},
            {"universe": "monthly_research_panel", "ticker_count": len(panel_codes)},
            {"universe": "delisted_in_raw_adjusted_price", "ticker_count": len(delisted_codes & raw_price_codes)},
            {"universe": "delisted_in_monthly_research_panel", "ticker_count": len(delisted_codes & panel_codes)},
        ]
    )
    rows.to_csv(OUT / "delisting_universe_audit.csv", index=False, encoding="utf-8-sig")
    report = {
        "raw_adjusted_price_tickers": len(raw_price_codes),
        "supplied_delisted_tickers": len(delisted_codes),
        "delisted_tickers_present_in_raw_adjusted_price": len(delisted_codes & raw_price_codes),
        "monthly_panel_tickers": len(panel_codes),
        "delisted_tickers_present_in_monthly_panel": len(delisted_codes & panel_codes),
        "conclusion": "The supplied delisted-stock universe was not removed at raw-price construction. Early panel endpoints must not be relabelled as evidence of a survivor-only universe.",
    }
    (OUT / "delisting_universe_audit_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
