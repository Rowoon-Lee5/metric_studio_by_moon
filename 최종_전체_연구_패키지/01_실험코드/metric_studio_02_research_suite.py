"""A falsification suite for the observable claims in Metric Studio 02.

This is a research register, not an alpha miner.  Every test has a fixed
signal, an explicit horizon, and an explicit statement of what cannot be
identified from price/volume/capitalisation data alone.
"""
from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = Path(r"C:\Users\benef\OneDrive\바탕 화면\직박구리2\직박구리\숭실대학교\(제출용) 금융인공지능및빅데이터응용_프로젝트_final\금융인공지능및빅데이터응용_프로젝트_final\raw data")
OUT = ROOT / "results"

def _monthly_mcap() -> pd.DataFrame:
    """Stream raw market-cap workbook and retain only month-end observations."""
    cache = OUT / "monthly_mcap_panel.pkl"
    if cache.exists(): return pd.read_pickle(cache)
    wb = load_workbook(RAW / "시가총액.xlsx", read_only=True, data_only=True); ws = wb.active
    codes = [str(x)[1:] if x else None for x in next(ws.iter_rows(min_row=9,max_row=9,values_only=True))[1:]]
    rows, last = [], None
    try:
        for row in ws.iter_rows(min_row=15, values_only=True):
            date = pd.Timestamp(row[0]); value = np.fromiter((float(x) if isinstance(x,(int,float)) else np.nan for x in row[1:]), dtype=float)
            current = (date, value)
            if last is not None and date.to_period("M") != last[0].to_period("M"): rows.append(last)
            last = current
        rows.append(last)
    finally: wb.close()
    result = pd.concat([pd.DataFrame({"date":d,"ticker":codes,"mcap":v}) for d,v in rows], ignore_index=True)
    result = result.dropna(subset=["ticker","mcap"]); result.to_pickle(cache)
    return result

def _cagr(r: pd.Series) -> float:
    r = r.dropna()
    return float((1 + r).prod() ** (12 / len(r)) - 1) if len(r) else np.nan

def _mdd(r: pd.Series) -> float:
    wealth=(1+r.fillna(0)).cumprod()
    return float((wealth/wealth.cummax()-1).min())

def _risk_stats(r: pd.Series) -> dict:
    return {"cagr":_cagr(r),"annualized_vol":float(r.std()*np.sqrt(12)),"mdd":_mdd(r),"months":int(r.notna().sum())}

def _quantile_forward(x: pd.DataFrame, signal: str, target: str, label: str, q: int=5) -> list[dict]:
    rows=[]
    for date, d in x.groupby("date"):
        d=d.dropna(subset=[signal,target])
        if len(d)<q*10: continue
        try: d=d.assign(bucket=pd.qcut(d[signal],q,labels=False,duplicates="drop"))
        except ValueError: continue
        for b,g in d.groupby("bucket"):
            rows.append({"test":label,"date":date,"bucket":int(b)+1,"mean_return":g[target].mean(),"n":len(g)})
    return rows

def main():
    price = pd.read_pickle(OUT / "monthly_price_adv_panel.pkl")
    x = price.merge(_monthly_mcap(), on=["date","ticker"], how="inner")
    x=x.sort_values(["ticker","date"]); x=x[(x.price>0)&(x.mcap>0)&(x.adv_21d>0)].copy()
    x["next_1m"] = x.groupby("ticker").price.shift(-1)/x.price-1
    x["next_12m"] = x.groupby("ticker").price.shift(-12)/x.price-1
    x["ret_1m"] = x.groupby("ticker").price.pct_change()
    x["ret_6m"] = x.groupby("ticker").price.pct_change(6)
    x["log_price"],x["log_mcap"] = np.log(x.price),np.log(x.mcap)
    x["turnover_proxy"] = x.adv_21d/x.mcap
    x["vol_12m"] = x.groupby("ticker").ret_1m.transform(lambda s:s.rolling(12,min_periods=8).std())
    rows=[]
    rows += _quantile_forward(x,"log_price","next_1m","price_level_vs_future_1m")
    rows += _quantile_forward(x,"log_mcap","next_1m","size_vs_future_1m")
    rows += _quantile_forward(x,"turnover_proxy","next_1m","turnover_proxy_vs_future_1m")
    rows += _quantile_forward(x,"ret_6m","next_12m","six_month_momentum_vs_future_12m")
    rows += _quantile_forward(x,"vol_12m","next_1m","low_volatility_vs_future_1m")
    # Shock persistence: an outlier is operationally informative only if its subsequent drift differs.
    x["cs_median"] = x.groupby("date").ret_1m.transform("median")
    x["cs_mad"] = x.groupby("date").ret_1m.transform(lambda s:(s-s.median()).abs().median())
    x["shock_z"] = .6745*(x.ret_1m-x.cs_median)/(x.cs_mad+1e-12)
    x["shock_group"] = np.select([x.shock_z<=-3,x.shock_z>=3],["negative_shock","positive_shock"],default="ordinary")
    for (group,liq),g in x.assign(liq=np.where(x.turnover_proxy>=x.groupby("date").turnover_proxy.transform("median"),"high_turnover","low_turnover")).groupby(["shock_group","liq"]):
        if len(g): rows.append({"test":"shock_persistence_1m","date":"ALL","bucket":f"{group}:{liq}","mean_return":g.next_1m.mean(),"n":len(g)})
    detail=pd.DataFrame(rows); detail.to_csv(OUT/"research_suite_detail.csv",index=False,encoding="utf-8-sig")
    summary=[]
    for test,g in detail[detail.date!="ALL"].groupby("test"):
        q=g.groupby("bucket").mean_return.mean()
        summary.append({"test":test,"bottom_bucket_mean_pct":float(q.iloc[0]*100),"top_bucket_mean_pct":float(q.iloc[-1]*100),"spread_pct":float((q.iloc[-1]-q.iloc[0])*100),"months":int(g.date.nunique())})
    shocks=detail[(detail.test=="shock_persistence_1m")].drop(columns=["test","date"])
    market=x.groupby("date").next_1m.mean(); market_stats={"monthly_up_probability":float((market>0).mean()),"monthly_return_skew":float(market.skew()),"monthly_excess_kurtosis":float(market.kurt()),"monthly_market_cagr":_cagr(market)}
    # Long-only hedge: a cash sleeve scales the equal-weight market to 12% target volatility.
    trailing_vol=market.rolling(12,min_periods=8).std()*np.sqrt(12)
    cash_weight=(.12/trailing_vol).clip(upper=1).fillna(1)
    hedge_stats={"fully_invested":_risk_stats(market),"cash_vol_target_12pct":_risk_stats(market*cash_weight),"meaning":"Cash is the hedge; no short position is used. This is a risk-overlay test, not an alpha test."}
    # Stop-loss counterfactual.  Monthly closes cannot test intramonth stops; the restriction is deliberate.
    stop_rows=[]
    for threshold in [.10,.20,.30]:
        outcomes=[]
        for _,g in x.groupby("ticker"):
            p=g.price.to_numpy(dtype=float)
            for i in range(len(p)-12):
                path=p[i+1:i+13]/p[i]-1
                hit=np.flatnonzero(path<=-threshold)
                outcomes.append(path[hit[0]] if len(hit) else path[-1])
        r=pd.Series(outcomes)
        stop_rows.append({"stop_threshold":threshold,"mean_12m_return":float(r.mean()),"median_12m_return":float(r.median()),"loss_probability":float((r<0).mean()),"observations":int(len(r))})
    baseline=[]
    for _,g in x.groupby("ticker"):
        p=g.price.to_numpy(dtype=float); baseline.extend(p[12:]/p[:-12]-1)
    stop_test={"monthly_close_buy_and_hold_12m":{"mean_12m_return":float(np.mean(baseline)),"median_12m_return":float(np.median(baseline)),"loss_probability":float(np.mean(np.array(baseline)<0)),"observations":len(baseline)},"stop_rules":stop_rows,"warning":"This is a market-wide holding-rule counterfactual, not a trading strategy. Intramonth barriers, transaction costs and re-entry are not observed."}
    report={"sample":{"start":str(x.date.min().date()),"end":str(x.date.max().date()),"observations":int(len(x)),"tickers":int(x.ticker.nunique())},"market_distribution":market_stats,"quintile_tests":summary,"shock_persistence":shocks.to_dict(orient="records"),"long_only_risk_overlay":hedge_stats,"stop_loss_counterfactual":stop_test,"identification_warning":"Turnover is not investor attention; this dataset contains neither search/news/ownership attention nor executable bid-ask spreads. Results are associations, not causal evidence."}
    (OUT/"research_suite_report.json").write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps(report,ensure_ascii=False,indent=2))
if __name__=="__main__": main()
