"""Audit an end-of-day stop rule on the selected small-cap holdings.

The supplied workbook has daily adjusted closes, but no intraday high/low,
quotes, or order records.  This deliberately does *not* claim to test a
trigger-price stop order.  A position is entered at the formation-date close;
when a later daily close first falls below the threshold, it exits at the next
available daily close and remains cash until the next monthly formation date.
This makes gap risk explicit and avoids filling a stop at a price never seen
in the data.
"""

from __future__ import annotations

import json
from bisect import bisect_left
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "results"
RAW = Path(r"C:\Users\benef\OneDrive\바탕 화면\직박구리2\직박구리\숭실대학교\(제출용) 금융인공지능및빅데이터응용_프로젝트_final\금융인공지능및빅데이터응용_프로젝트_final\raw data")
THRESHOLDS = [0.10, 0.20, 0.30]


def load_selected_daily_prices(tickers: set[str]) -> tuple[list[pd.Timestamp], dict[str, np.ndarray]]:
    """Stream only selected columns from the large daily adjusted-price panel."""
    workbook = load_workbook(RAW / "수정주가.xlsx", read_only=True, data_only=True)
    sheet = workbook.active
    codes = [str(value)[1:] if value else None for value in next(sheet.iter_rows(min_row=9, max_row=9, values_only=True))[1:]]
    positions = {code: index for index, code in enumerate(codes) if code in tickers}
    prices = {code: [] for code in positions}
    dates: list[pd.Timestamp] = []
    try:
        for row in sheet.iter_rows(min_row=15, values_only=True):
            dates.append(pd.Timestamp(row[0]).normalize())
            for code, index in positions.items():
                value = row[index + 1]
                prices[code].append(float(value) if isinstance(value, (int, float)) and value > 0 else np.nan)
    finally:
        workbook.close()
    return dates, {code: np.asarray(values, dtype=float) for code, values in prices.items()}


def metric(returns: pd.Series) -> dict[str, float]:
    returns = returns.dropna()
    wealth = (1 + returns).cumprod()
    t_stat = float(returns.mean() / (returns.std(ddof=1) / np.sqrt(len(returns)))) if returns.std(ddof=1) > 0 else 0.0
    return {
        "months": int(len(returns)),
        "cagr": float(wealth.iloc[-1] ** (12 / len(returns)) - 1),
        "mdd": float((wealth / wealth.cummax() - 1).min()),
        "t_stat": t_stat,
        "mean_monthly_return": float(returns.mean()),
    }


def main() -> None:
    holdings = pd.read_csv(OUT / "smallcap_selected_holdings.csv", dtype={"ticker": str}, parse_dates=["date"])
    holdings["date"] = holdings.date.dt.normalize()
    formation_dates = sorted(holdings.date.unique())
    dates, prices = load_selected_daily_prices(set(holdings.ticker))
    date_index = {date: index for index, date in enumerate(dates)}
    rows = []
    arms = []
    for formation, next_formation in zip(formation_dates[:-1], formation_dates[1:]):
        start = date_index.get(pd.Timestamp(formation))
        end = date_index.get(pd.Timestamp(next_formation))
        if start is None or end is None or end <= start:
            continue
        month = holdings[holdings.date == formation]
        values: dict[str, list[float]] = {"hold": []} | {f"stop_{int(t * 100)}": [] for t in THRESHOLDS}
        for ticker in month.ticker:
            series = prices.get(ticker)
            if series is None or not np.isfinite(series[start]) or not np.isfinite(series[end]):
                continue
            entry, final = series[start], series[end]
            values["hold"].append(final / entry - 1)
            for threshold in THRESHOLDS:
                arm = f"stop_{int(threshold * 100)}"
                trigger = None
                for index in range(start + 1, end + 1):
                    if np.isfinite(series[index]) and series[index] / entry - 1 <= -threshold:
                        trigger = index
                        break
                if trigger is None:
                    values[arm].append(final / entry - 1)
                    arms.append({"formation_date": formation, "ticker": ticker, "threshold": threshold, "triggered": False, "stop_return": final / entry - 1, "hold_return": final / entry - 1})
                    continue
                exit_index = min(trigger + 1, end)
                while exit_index <= end and not np.isfinite(series[exit_index]):
                    exit_index += 1
                exit_price = series[exit_index] if exit_index <= end else final
                stop_return = exit_price / entry - 1
                values[arm].append(stop_return)
                arms.append({"formation_date": formation, "ticker": ticker, "threshold": threshold, "triggered": True, "stop_return": stop_return, "hold_return": final / entry - 1})
        if not values["hold"]:
            continue
        for arm, returns in values.items():
            rows.append({"date": formation, "rule": arm, "monthly_return": float(np.mean(returns)), "observed_holdings": len(returns)})
    monthly = pd.DataFrame(rows)
    monthly.to_csv(OUT / "stop_loss_daily_close_monthly.csv", index=False, encoding="utf-8-sig")
    arm_data = pd.DataFrame(arms)
    arm_data.to_csv(OUT / "stop_loss_daily_close_events.csv", index=False, encoding="utf-8-sig")
    summary = []
    for rule, subset in monthly.groupby("rule"):
        result = metric(subset.monthly_return)
        if rule != "hold":
            threshold = int(rule.split("_")[1]) / 100
            events = arm_data[arm_data.threshold == threshold]
            triggered = events[events.triggered]
            result.update(
                {
                    "rule": rule,
                    "threshold": threshold,
                    "stop_hit_rate": float(events.triggered.mean()),
                    "triggered_positions": int(len(triggered)),
                    "observed_positions": int(len(events)),
                    "recovered_by_month_end_rate": float((triggered.hold_return > triggered.stop_return).mean()) if len(triggered) else 0.0,
                }
            )
        else:
            result.update({"rule": rule, "threshold": None, "stop_hit_rate": None, "triggered_positions": None, "observed_positions": int(len(arm_data) / len(THRESHOLDS)), "recovered_by_month_end_rate": None})
        summary.append(result)
    summary_df = pd.DataFrame(summary).sort_values("threshold", na_position="first")
    summary_df.to_csv(OUT / "stop_loss_daily_close_summary.csv", index=False, encoding="utf-8-sig")
    report = {
        "rule": "At a daily close below the loss threshold, exit at the next available daily close; cash is held until the next monthly formation date.",
        "comparison": "Same selected small-cap holdings, equal-weighted across positions with both formation and next-formation adjusted closes observed.",
        "summary": summary_df.to_dict(orient="records"),
        "limits": [
            "This is an end-of-day adjusted-close proxy, not an intraday stop-order backtest.",
            "The raw panel has no intraday low/high, bid-ask spread, order-book depth, or realised fill records.",
            "The conditional sample excludes holdings without both endpoint adjusted closes and therefore does not settle terminal-event cash values.",
            "Transaction costs and capacity are held outside this overlay comparison; it tests the timing rule, not executable net alpha.",
        ],
    }
    (OUT / "stop_loss_daily_close_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
