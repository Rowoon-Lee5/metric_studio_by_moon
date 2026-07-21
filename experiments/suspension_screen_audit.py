"""Apply the supplied Dataguide trading-suspension panel to the leading small-cap node.

This is a real implementable correction: a stock labelled non-normal on the
formation date is excluded before selection. It does not fabricate delisting
returns; those require a dedicated event/return source.
"""
from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "results"
RAW = Path(r"C:\Users\benef\OneDrive\바탕 화면\직박구리2\직박구리\숭실대학교\(제출용) 금융인공지능및빅데이터응용_프로젝트_final\금융인공지능및빅데이터응용_프로젝트_final\raw data")
SUSPENSION = RAW / "거래정지구분.xlsx"
PARTICIPATION = 0.05


def cagr(series: pd.Series) -> float:
    return float((1 + series).prod() ** (12 / len(series)) - 1)


def t_stat(series: pd.Series) -> float:
    return float(series.mean() / (series.std(ddof=1) / np.sqrt(len(series))))


def mdd(series: pd.Series) -> float:
    wealth = (1 + series).cumprod()
    return float((wealth / wealth.cummax() - 1).min())


def read_monthly_suspensions(wanted_dates: set[pd.Timestamp], wanted_tickers: set[str]) -> pd.DataFrame:
    wb = load_workbook(SUSPENSION, read_only=True, data_only=True)
    ws = wb.active
    codes_row = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    codes = [str(v).replace("A", "").zfill(6) if v else None for v in codes_row[1:]]
    keep_cols = [(i + 1, code) for i, code in enumerate(codes) if code in wanted_tickers]
    rows = []
    for values in ws.iter_rows(min_row=15, values_only=True):
        raw_date = values[0]
        if not isinstance(raw_date, (pd.Timestamp,)):
            raw_date = pd.Timestamp(raw_date)
        date = pd.Timestamp(raw_date).normalize()
        if date not in wanted_dates:
            continue
        for column, ticker in keep_cols:
            status = values[column]
            if status and str(status).strip() != "정상":
                rows.append({"date": date, "ticker": ticker, "suspension_status": str(status).strip()})
    wb.close()
    return pd.DataFrame(rows)


def main() -> None:
    price = pd.read_pickle(OUT / "monthly_price_adv_panel.pkl").sort_values(["ticker", "date"]).copy()
    mcap = pd.read_pickle(OUT / "monthly_mcap_panel.pkl")
    price["ticker"] = price.ticker.astype(str).str.zfill(6)
    wanted_dates = set(pd.to_datetime(price.date).dt.normalize().unique())
    suspended = read_monthly_suspensions(wanted_dates, set(price.ticker.unique()))
    suspended.to_csv(OUT / "monthly_suspension_flags.csv", index=False, encoding="utf-8-sig")

    x = price.merge(mcap, on=["date", "ticker"], how="left")
    x["next"] = x.groupby("ticker").price.shift(-1) / x.price - 1
    x = x.merge(suspended.assign(suspended=True), on=["date", "ticker"], how="left")
    x["suspended"] = x.suspended.fillna(False).astype(bool)
    rows = []
    for date, day in x.groupby("date"):
        # A future missing price cannot be a formation-time exclusion rule.
        # Endpoint sensitivity is handled separately in closure_audits.py.
        day = day.dropna(subset=["adv_21d", "mcap"])
        day = day[(day.adv_21d > 0) & (day.price > 0) & (day.mcap > 0)]
        base = day.nlargest(max(50, int(len(day) * 0.9)), "adv_21d").nsmallest(50, "mcap")
        screened = day.loc[~day.suspended].nlargest(max(50, int((~day.suspended).sum() * 0.9)), "adv_21d").nsmallest(50, "mcap")
        for label, selected in {"unfiltered": base, "suspension_screened": screened}.items():
            if len(selected) < 50:
                continue
            aum = 1e8
            filled = np.minimum(aum / 50, PARTICIPATION * selected.adv_21d.to_numpy())
            ratio = np.divide(filled, selected.adv_21d.to_numpy(), out=np.zeros(50), where=selected.adv_21d.to_numpy() > 0)
            one_way = 0.001 + 0.005 * 0.5 * np.sqrt(ratio)
            net = (filled * selected.next.fillna(0.0).to_numpy()).sum() / aum - 2 * (filled * one_way).sum() / aum
            rows.append({"date": date, "variant": label, "net_return": net, "holdings": len(selected), "excluded_suspended": int(day.suspended.sum())})
    returns = pd.DataFrame(rows)
    summary = returns.groupby("variant").net_return.agg(months="size", mean_return="mean", cagr=cagr, t_stat=t_stat, mdd=mdd).reset_index()
    summary.to_csv(OUT / "suspension_screened_smallcap_summary.csv", index=False, encoding="utf-8-sig")
    returns.to_csv(OUT / "suspension_screened_smallcap_returns.csv", index=False, encoding="utf-8-sig")
    report = {
        "screen": "At formation, exclude any stock whose supplied Dataguide status is not 정상.",
        "suspension_flag_rows": int(len(suspended)),
        "flagged_tickers": int(suspended.ticker.nunique()) if len(suspended) else 0,
        "strategy": "small_cap|liquidity_top_90pct|50_holdings|KRW100m|cost_0.5x",
        "summary": summary.to_dict(orient="records"),
    }
    (OUT / "suspension_screen_audit_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
