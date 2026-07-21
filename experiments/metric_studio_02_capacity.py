"""Metric Studio 02 experiment: liquidity as a capacity constraint.

Raw-data source is deliberately explicit.  The workbook panels are streamed
because each has about 4,000 tickers x 6,500 days.
"""
from __future__ import annotations

import json
from collections import deque
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = Path(r"C:\Users\benef\OneDrive\바탕 화면\직박구리2\직박구리\숭실대학교\(제출용) 금융인공지능및빅데이터응용_프로젝트_final\금융인공지능및빅데이터응용_프로젝트_final\raw data")
OUT = ROOT / "results"
N_HOLDINGS, UNIVERSE_PCTS = 30, np.arange(.2, 1.01, .1)
AUMS = [100_000_000, 1_000_000_000, 10_000_000_000]
BASE_COST, IMPACT = .001, .005  # one-way: 10bp + 50bp*sqrt(order/ADV)

def open_panel(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True); ws = wb.active
    codes = [str(x)[1:] if x else None for x in next(ws.iter_rows(min_row=9, max_row=9, values_only=True))[1:]]
    names = list(next(ws.iter_rows(min_row=10, max_row=10, values_only=True))[1:])
    return wb, codes, names, ws.iter_rows(min_row=15, values_only=True)

def floats(row):
    return np.fromiter((float(x) if isinstance(x, (int, float)) else np.nan for x in row[1:]), dtype=float)

def make_monthly_panel():
    pwb, codes, names, prows = open_panel(RAW / "수정주가.xlsx")
    vwb, vcodes, _, vrows = open_panel(RAW / "거래량.xlsx")
    positions = {c:i for i,c in enumerate(vcodes) if c}
    index = np.array([positions.get(c, -1) for c in codes])
    if (index < 0).mean() > .05: raise ValueError("price-volume ticker match failure")
    saved, window, previous = [], deque(maxlen=21), None
    try:
        for pr, vr in zip(prows, vrows):
            date = pd.Timestamp(pr[0])
            if date != pd.Timestamp(vr[0]): raise ValueError("date alignment failure")
            price, raw_volume = floats(pr), floats(vr)
            volume = np.full(len(codes), np.nan); valid = index >= 0
            volume[valid] = raw_volume[index[valid]]
            window.append(price * volume)
            current = (date, price.copy(), np.nanmean(np.vstack(window), axis=0))
            if previous is not None and date.to_period("M") != previous[0].to_period("M"): saved.append(previous)
            previous = current
        if previous is not None: saved.append(previous)
    finally:
        pwb.close(); vwb.close()
    frames = [pd.DataFrame({"date":d,"ticker":codes,"name":names,"price":p,"adv_21d":a}) for d,p,a in saved]
    return pd.concat(frames, ignore_index=True).dropna(subset=["ticker","price","adv_21d"])

def evaluate(panel):
    x = panel.sort_values(["ticker","date"]).copy()
    x["next_return"] = x.groupby("ticker").price.shift(-1) / x.price - 1
    x["ret_1m"] = x.groupby("ticker").price.pct_change()
    rows=[]
    for date, d in x.groupby("date"):
        d=d.dropna(subset=["ret_1m","next_return","adv_21d"]); d=d[(d.price>0)&(d.adv_21d>0)]
        for pct in UNIVERSE_PCTS:
            liquid=d.nlargest(max(N_HOLDINGS,int(len(d)*pct)),"adv_21d")
            if len(liquid)<N_HOLDINGS: continue
            selected=liquid.nsmallest(N_HOLDINGS,"ret_1m"); gross=selected.next_return.mean(); adv=selected.adv_21d.mean()
            for aum in AUMS:
                cost=BASE_COST+IMPACT*np.sqrt((aum/N_HOLDINGS)/adv)
                rows.append(dict(date=date,universe_pct=pct,aum_krw=aum,gross_return=gross,net_return=gross-min(.30,2*cost),one_way_cost=cost,mean_adv_krw=adv,n_universe=len(liquid)))
    monthly=pd.DataFrame(rows)
    cagr=lambda r:(1+r).prod()**(12/len(r))-1
    summary=monthly.groupby(["universe_pct","aum_krw"]).agg(months=("net_return","size"),gross_cagr=("gross_return",cagr),net_cagr=("net_return",cagr),avg_one_way_cost=("one_way_cost","mean"),avg_adv=("mean_adv_krw","mean")).reset_index()
    return monthly,summary

def main():
    OUT.mkdir(exist_ok=True); cache=OUT/"monthly_price_adv_panel.pkl"
    panel=pd.read_pickle(cache) if cache.exists() else make_monthly_panel()
    if not cache.exists(): panel.to_pickle(cache)
    monthly,summary=evaluate(panel)
    monthly.to_csv(OUT/"monthly_capacity_returns.csv",index=False,encoding="utf-8-sig")
    summary.to_csv(OUT/"capacity_summary.csv",index=False,encoding="utf-8-sig")
    best=summary.loc[summary.groupby("aum_krw").net_cagr.idxmax()].sort_values("aum_krw")
    report={"sample":{"start":str(panel.date.min().date()),"end":str(panel.date.max().date()),"monthly_snapshots":int(panel.date.nunique()),"tickers":int(panel.ticker.nunique())},"assumptions":{"signal":"monthly 1-month reversal; bottom 30","cost":"10bp + 50bp*sqrt(order/21-day ADV), one-way","turnover":"100% monthly"},"net_cagr_optimum_by_aum":best.to_dict(orient="records")}
    (OUT/"research_report.json").write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps(report,ensure_ascii=False,indent=2))
if __name__=="__main__": main()
