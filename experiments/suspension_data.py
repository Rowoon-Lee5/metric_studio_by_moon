"""Read the supplied DataGuide trading-suspension panel at formation dates.

This module intentionally marks only information available at each formation
date.  It must not remove a stock merely because the stock is delisted later.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


RAW = Path(r"C:\Users\benef\OneDrive\바탕 화면\직박구리2\직박구리\숭실대학교\(제출용) 금융인공지능및빅데이터응용_프로젝트_final\금융인공지능및빅데이터응용_프로젝트_final\raw data")
SUSPENSION = RAW / "거래정지구분.xlsx"


def read_monthly_suspensions(panel: pd.DataFrame) -> pd.DataFrame:
    """Return non-normal statuses for the panel's own dates and tickers."""
    if not SUSPENSION.exists():
        raise FileNotFoundError(f"거래정지구분 파일을 찾지 못했습니다: {SUSPENSION}")

    wanted_dates = set(pd.to_datetime(panel["date"]).dt.normalize().unique())
    wanted_tickers = set(panel["ticker"].astype(str).str.zfill(6).unique())
    wb = load_workbook(SUSPENSION, read_only=True, data_only=True)
    ws = wb.active
    codes_row = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    codes = [str(v).replace("A", "").zfill(6) if v else None for v in codes_row[1:]]
    keep_cols = [(i + 1, code) for i, code in enumerate(codes) if code in wanted_tickers]
    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=15, values_only=True):
        raw_date = values[0]
        if raw_date is None:
            continue
        date = pd.Timestamp(raw_date).normalize()
        if date not in wanted_dates:
            continue
        for column, ticker in keep_cols:
            status = values[column]
            if status and str(status).strip() != "정상":
                rows.append({"date": date, "ticker": ticker, "suspension_status": str(status).strip()})
    wb.close()
    return pd.DataFrame(rows, columns=["date", "ticker", "suspension_status"])
